# -*- coding: utf-8 -*-
"""
Created on Thu Sep 26 14:56:31 2024

@author: shivanshu.mall


Bhandara_Bank_1.0.6.py:04-06-2025 Changing the Embo input file Name 
Bhandara_Bank_1.0.5.py:05-03-2025 Fixxing the issue in the label file 
Bhandara_Bank_1.0.4.py:04-03-2025 Adding the barcode in packing list and changing the file structure


"""

import warnings
from openpyxl.styles import PatternFill, Border, Side, Alignment
import barcode
from barcode.writer import ImageWriter
from io import BytesIO
from PIL import Image
import pandas as pd
import glob
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
import pandas as pd
import time
from datetime import date
from datetime import datetime
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
import collections
import openpyxl
from fpdf import FPDF
from prettytable import PrettyTable
import time as tmt
import shutil
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
embo_flag= 0
address_flag = 0
current_date = datetime.now()
date_1=current_date.strftime("%Y-%m-%d")
sn_0 = 0
total_qty = 0
count = 0
ts = date.today()
x = (ts)
Time = time.localtime()
fileqty = []
read_only_flag = ''
def merge_address_sort(line):     
    parts = line.strip().split(',') 
          
    return parts[8],parts[10]

def extract_info_Address(line):     
    parts = line.strip().split('|') 
          
    return parts[7],parts[9]

def extract_info_embo(line):     
    parts = line.strip().split('|')       
    return parts[-2],parts[-1]

def Excel_converter(file):
    try:
        df = pd.read_csv(file, sep='|', dtype=object)
        df.to_excel(file[:-4]+'.xlsx', 'Sheet1',index=False)            
    except Exception as e:
        print('Error in Excel Converter',e)

def split_files(f):
    file_count = 0
    with open(f, 'r') as infile:
        lines = infile.readlines()
        size = len(lines)
        if size <= 500:
            file_count += 1
            with open(f[:-4] + '_' + str(file_count) + '.txt', 'w') as outfile:
                outfile.writelines(lines)
        else:
            for i in range(0, size, 500):
                file_count += 1
                chunk = lines[i:i + 500]
                with open(f[:-4] + '_' + str(file_count) + '.txt', 'w') as outfile:
                    outfile.writelines(chunk)
            
            
def convert_to_read_only(file_path):
    password = 'your_password_here' 
    global read_only_flag
    
    if read_only_flag == 'NO':
        return
    else:
        try:
            wb = load_workbook(file_path)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                ws.protection.sheet = True
                ws.protection.password = password
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                ws.sheet_view.zoomScale = 85
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for col in range(1, ws.max_column + 1):
                    column_letter = get_column_letter(col)
                    max_length = 0
                    for cell in ws[column_letter]:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = (max_length + 2) * 1.2
            
            wb.save(file_path)
        except Exception as e:
            print(f"Error converting '{file_path}': {e}")   
            
  
            
def label(file,total_qty,i):
    global sn_0
    bin_ = file.split('_')[3].strip()
    with open(file,'r') as infile, open('xxBHANDARA_LABEL_FILE.csv','a') as lbl_out, open('xxBHANDARA_SCANNING_FILE.csv','a') as SCAN_out, open('xxBHANDARA_LABEL_SCANNING_FILE.csv','a') as FPA_out,open('xxBHANDARA_PACKAGING_FILE.csv','a') as packaging_outfile:
        contents = infile.readlines()
        size = len(contents)
        ln = contents[0]
        value_1 = ln.split('|')
        sn_0 += 1 
        Name = value_1[0].rstrip()
        Address1 = value_1[1].rstrip()
        Address2 = value_1[2].rstrip()
        Address3 = value_1[3].rstrip()
        City = value_1[4].rstrip()
        State = value_1[5].rstrip()
        Pin = value_1[6].rstrip()
        Branch_ID = value_1[7].rstrip()
        Branch_Name = value_1[8].rstrip()
        Account_No =value_1[9].rstrip()
        Cust_Card_No = value_1[10].rstrip()
        mask_card_number=Cust_Card_No[:4]+'XXXX XXXX'+Cust_Card_No[-4:]
        first_six_bin = Cust_Card_No[:7].replace(' ','')
        Expiry_Date = value_1[11].rstrip()
        first_Card_Id = value_1[12].rstrip()
        Sub_Type = value_1[13].rstrip()
        Mobile_No = value_1[14].rstrip()
        filename = value_1[-1].rstrip()
        
        last_card_id = contents[len(contents) - 1].split('|')[12]
        lbl_out.write(Branch_Name+'|'+Branch_ID.zfill(4)+'|'+first_Card_Id+'|'+last_card_id+'|'+str(size)+'|'+Branch_ID.zfill(4)+' - '+str(i).zfill(3)+'/'+str(total_qty).zfill(3)+'|'+filename+'|'+''+'|'+''+'|'+Name+'|'+Address1+'|'+Address2+'|'+City+'|'+State+'|'+Pin+'|'+Mobile_No+'|'+''+'|'+''+'\n')
        for line in contents:
            data = line.split('|')
            sn_0 += 1 
            Name = data[0].rstrip()
            Address1 = data[1].rstrip()
            Address2 = data[2].rstrip()
            Address3 = data[3].rstrip()
            City = data[4].rstrip()
            State = data[5].rstrip()
            Pin = data[6].rstrip()
            Branch_ID = data[7].rstrip()
            Branch_Name = data[8].rstrip()
            Account_No =data[9].rstrip()
            Cust_Card_No = data[10].rstrip()
            mask_card_number=Cust_Card_No[:4]+'XXXX XXXX'+Cust_Card_No[-4:]
            first_six_bin = Cust_Card_No[:7]
            Expiry_Date = data[11].rstrip()
            first_Card_Id = data[12].rstrip()
            Sub_Type = data[13].rstrip()
            Mobile_No = data[14].rstrip()
            filename = data[-1].rstrip()
            
            SCAN_out.write(first_Card_Id+'|'+mask_card_number+'|'+'Bluedart'+'|'+''+'|'+Branch_ID.zfill(4)+' - '+str(i).zfill(3)+'/'+str(total_qty).zfill(3)+'|'+Branch_ID.zfill(4)+'|'+bin_+'\n')
            FPA_out.write(first_Card_Id+'|'+mask_card_number+'|'+'Bluedart'+'|'+''+'|'+Branch_ID.zfill(4)+' - '+str(i).zfill(3)+'/'+str(total_qty).zfill(3)+'|'+Branch_ID.zfill(4)+'\n')
            packaging_outfile.write(first_Card_Id+'|'+mask_card_number+'|'+Branch_ID.zfill(4) +'|'+Account_No+'|'+Branch_Name+'|'+str(i).zfill(3)+'/'+str(total_qty).zfill(3)+'\n')




class PDF(FPDF):
    def __init__(self):
        super().__init__()
        self.total_pages = 0

    def header(self):
        pass

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'B', 7)
        current_page = str(self.page_no()).zfill(3)  
        total_pages = str(self.total_pages).zfill(3) 
        self.cell(0, 10, f'Page {current_page} of {total_pages}', align='C')



def add_barcode_to_pdf(pdf, value):
    try:
        global count
        count += 1
        code128 = barcode.get_barcode_class("code128")
        barcode_obj = code128(value, writer=ImageWriter())
        barcode_buffer = BytesIO()
        barcode_obj.write(barcode_buffer, {'module_width': 0.3, 'module_height': 15.0})
        barcode_buffer.seek(0)
        barcode_image = Image.open(barcode_buffer)
        temp_path = f"temp_barcode_{count}.png"
        barcode_image.save(temp_path, format="PNG")
        pdf.image(temp_path, x=(250 - 100) / 2, y=3, w=65, h=17)
        os.remove(temp_path)

    except Exception as e:
        print(f"Error generating barcode for value {value}: {e}")

def generate_pdf_from_excel(file_path, output_pdf):
    """Read Excel file and generate a PDF with barcodes and data table."""

 
    df = pd.read_excel(file_path, dtype=str)

    if df.empty:
        print("The Excel file is empty. Exiting.")
        return

    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.set_font("Arial", size=8)
    pdf.set_left_margin(10)
    pdf.set_right_margin(10)
    pdf.set_top_margin(20)

    row_height = 5.05
    rows_per_page = 50

    previous_branch_code = None
    previous_box = None

    for index, row in df.iterrows():
        current_branch_code = row["Branch ID"]
        current_box = row["Box No."]
        if current_branch_code != previous_branch_code or current_box != previous_box:
            pdf.add_page() 
            combined_value = f"* {current_branch_code} - {current_box} - P *"
            add_barcode_to_pdf(pdf, combined_value)

            # Add table headers
            pdf.set_font("Arial", "B", 8)
            headers = df.columns.tolist()
            pdf.set_fill_color(200, 200, 200)
            cell_widths = [12, 20, 30, 17, 25, 25,15]

            for header, width in zip(headers, cell_widths):
                pdf.cell(width, row_height, header, border=1, align="C", fill=True)
            pdf.ln(row_height)

        pdf.set_font("Arial", "", 8)
        for value, width in zip(row.tolist(), cell_widths):
            pdf.cell(width, row_height, str(value), border=1, align="C")
        pdf.ln(row_height)

        previous_branch_code = current_branch_code
        previous_box = current_box

    pdf.total_pages = pdf.page_no()

    pdf2 = PDF()
    pdf2.total_pages = pdf.total_pages
    pdf2.set_auto_page_break(auto=True, margin=10)
    pdf2.set_font("Arial", size=8)
    pdf2.set_left_margin(10)
    pdf2.set_right_margin(10)
    pdf2.set_top_margin(20)

    previous_branch_code = None
    previous_box = None

    for index, row in df.iterrows():
        current_branch_code = row["Branch ID"]
        current_box = row["Box No."]

        if current_branch_code != previous_branch_code or current_box != previous_box:
            pdf2.add_page()  # New page

            # Generate barcode
            combined_value = f"* {current_branch_code} - {current_box} - P *"
            add_barcode_to_pdf(pdf2, combined_value)

            # Add table headers
            pdf2.set_font("Arial", "B", 8)
            headers = df.columns.tolist()
            pdf2.set_fill_color(200, 200, 200)
            cell_widths = [12, 20, 30, 17, 25, 25,15]

            for header, width in zip(headers, cell_widths):
                pdf2.cell(width, row_height, header, border=1, align="C", fill=True)
            pdf2.ln(row_height)

        # Add row data
        pdf2.set_font("Arial", "", 8)
        for value, width in zip(row.tolist(), cell_widths):
            pdf2.cell(width, row_height, str(value), border=1, align="C")
        pdf2.ln(row_height)

        previous_branch_code = current_branch_code
        previous_box = current_box

    # Save final PDF
    pdf2.output(output_pdf)
    print(f"\nPDF generated successfully: {output_pdf}")

           

input_files = glob.glob('TBDCCB*.csv') + glob.glob('*-DCCB*.txt')
if input_files == []:
    print('No input file present')
else:
    for f in input_files:
        with open(f,'r') as infile:
            if f.__contains__('TBDCCB'):
                
                with open(f,'r') as add_infile, open('xBDCCB_Merge_Address_file.csv','a') as add_outfile:
                    add_file_contents = add_infile.readlines()
                    fileqty.append(f)
                    occurrences = collections.Counter(fileqty)
                    for line in add_file_contents[1:]:
                        add_outfile.write(line.replace('\n','')+','+f+'\n')
                        
                
                
            elif f.__contains__('-DCCB'):
                with open(f,'r') as embo_infile, open('xBDCCB_Merge_EMBO_file.csv','a') as Embo_outfile:
                    fileqty.append(f)
                    occurrences = collections.Counter(fileqty)
                    embo_file_contents = embo_infile.readlines()
                    Embo_outfile.writelines(embo_file_contents)
    
    
    
    
    with open('xBDCCB_Merge_Address_file.csv','r') as mrg_add_file,open('xBDCCB_Merge_EMBO_file.csv','r') as mrg_emb_file, open('BDCCB_Merge_EMBO_file.txt','a') as mrg_emb_outfile:
        address_content = mrg_add_file.readlines()
        emb_contents = mrg_emb_file.readlines()
        for line in address_content:
            value = line.split(',')
            branch_id = value[8]
            branch_acc_no = value[10]
            card_number = value[11]
            for ln in emb_contents:
                card_no = ln[7:26].rstrip()
                if card_number == card_no:
                    mrg_emb_outfile.write(ln.replace('\n','')+'|'+branch_id.zfill(4)+'|'+branch_acc_no+'\n')
    
    os.remove('xBDCCB_Merge_EMBO_file.csv')
    
    with open('BDCCB_Merge_EMBO_file.txt','r') as mrg_emb_infile, open('BDCCB_Merge_EMBO_sortred_file.csv','a') as emb_outfile:
        contents = mrg_emb_infile.readlines()
        sorted_lines = sorted(contents, key=extract_info_embo)
        total_qty_lot = len(sorted_lines)
        emb_outfile.writelines(sorted_lines)
    
    with open('xBDCCB_Merge_Address_file.csv','r') as mrg_emb_infile, open('BDCCB_Merge_Address_sortred_file.csv','a') as emb_outfile:
        contents = mrg_emb_infile.readlines()
        sorted_lines = sorted(contents, key=merge_address_sort)
        emb_outfile.writelines(sorted_lines)
    
    
    with open('BDCCB_Merge_EMBO_sortred_file.csv','r') as emb_infile:
       contents =  emb_infile.readlines()
       size = len(contents)
       
       for line in contents:
           card_number = line.split(',')[1].strip()
           bin_num = line[7:14].replace(' ', '')
           filename = f'xBHANDARA_EMBO_FILE_{bin_num}_{date_1}.txt'
           with open(filename,'a') as emb_out:
                emb_out.write(line)


    for f in glob.glob('xBHANDARA_EMBO_FILE_*.txt'):
        with open(f,'r') as infile:
            contents = infile.readlines()
            size = len(contents)
            with open(f[1:-4]+'_'+str(size).zfill(4)+'.txtt','a') as outfile:
                outfile.writelines(contents)
                
                
            



    os.remove('BDCCB_Merge_EMBO_sortred_file.csv')
        
    
    
    with open('BDCCB_Merge_Address_sortred_file.csv','r') as add_infile:
        contents = add_infile.readlines()
        for line in contents:
            value = line.split(',')
            Name = value[1].rstrip()
            Address1 = value[2].rstrip()
            Address2 = value[3].rstrip()
            Address3 = value[4].rstrip()
            City = value[5].rstrip()
            State = value[6].rstrip()
            Pin = value[7].rstrip()
            Branch_ID = value[8].rstrip()
            Branch_Name = value[9].rstrip()
            Account_No =value[10].rstrip()
            Cust_Card_No = value[11].rstrip()
            mask_card_number=Cust_Card_No[:4]+'XXXX XXXX'+Cust_Card_No[-4:]
            first_six_bin = Cust_Card_No[:7].replace(' ','')
            Expiry_Date = value[12].rstrip()
            Card_Id = value[13].rstrip()
            Sub_Type = value[14].rstrip()
            Mobile_No = value[15].rstrip()
            filename = value[-1].rstrip()
            
            
            with open(f'BHANDARA_BATCH_ID_{first_six_bin}_{Branch_ID.zfill(4)}_FILE.txt','a') as ffoutfile:
                ffoutfile.write(Name+'|'+Address1+'|'+Address2+'|'+Address3+'|'+City+'|'+State+'|'+Pin+'|'+Branch_ID.zfill(4)+'|'+Branch_Name+'|'+Account_No+'|'+mask_card_number+'|'+Expiry_Date+'|'+Card_Id+'|'+Sub_Type+'|'+Mobile_No+'|'+filename+'\n')
            
    input_files = glob.glob('BHANDARA_BATCH_ID_*.txt') 
    # print(input_files)
    i = 0
    total_qty_box = 0
    for file in input_files:
        with open(file,'r') as infile:#,open('xxBHANDARA_PACKAGING_FILE.csv','a') as packaging_outfile:
            bin_num = file.split('_')[3].strip()
            contents = infile.readlines()
            size = len(contents)
            for line in contents:
                data = line.split('|')
                Name = data[0].rstrip()
                Branch_ID = data[7].rstrip()
                Branch_Name = data[8].rstrip()
                Account_No = data[9]
                Cust_Card_No =data[10]
                Card_Id = data[12]
                with open(f'xxBHANDARA_FF_MIS_FILE_{bin_num}.csv','a') as outfile,open(f'xxBHANDARA_MIS_FILE_{bin_num}.csv','a') as mis_outfile:
                    outfile.write(line.replace('\n','')+'|'+str(size)+'\n')
                    mis_outfile.write(line)
                # packaging_outfile.write(Name+'|'+Cust_Card_No+'|'+Branch_Name+'|'+Account_No+'|'+Branch_ID.zfill(2)+'|'+Card_Id+'\n')
    for fl in glob.glob('xxBHANDARA_FF_MIS_FILE_*'): 
        sr_no = 0    
        with open(fl,'r') as infile:
            bin_num = fl.split('_')[4][:-4].strip()
            contents = infile.readlines()
            size = len(contents)
            with open(f'BHANDARA_FF_MIS_FILE_{bin_num}_{date_1}_{str(size).zfill(4)}.csv','a') as ff_out:
                ff_out.write('Sr No.|Name|Address1|Address2|Address3|City|State|Pin|Branch ID|Branch Name|Account No.|Cust Card No.|Expiry Date|Card Id.|Sub Type|Mobile No.|File Name|Branch Qty.\n')
                for line in contents:
                    sr_no +=1 
                    ff_out.write(str(sr_no).zfill(5)+'|'+line)   

    # for file in glob.glob('BHANDARA_FF_MIS_FILE_*.txt'):
    for fl in glob.glob('xxBHANDARA_MIS_FILE*'):        
        with open(fl,'r') as infile:
            bin_num = fl.split('_')[3][:-4].strip()
            contents = infile.readlines()
            size = len(contents)
            sr_no = 0
            with open(f'BHANDARA_MIS_FILE_{bin_num}_{date_1}_{str(size).zfill(4)}.csv','a') as mis_out:
                mis_out.write('Sr No.|Name|Address1|Address2|Address3|City|State|Pin|Branch ID|Branch Name|Account No.|Cust Card No.|Expiry Date|Card Id.|Sub Type|Mobile No.|File Name\n')
                for line in contents:
                    sr_no += 1 
                    mis_out.write(str(sr_no).zfill(5)+'|'+line) 
    
    excel_inputs = glob.glob('BHANDARA_MIS_FILE_*.csv') +glob.glob('BHANDARA_FF_MIS_FILE_*.csv')
    for xl in excel_inputs:
        Excel_converter(xl)
        os.remove(xl)
    
    
    os.remove('BDCCB_Merge_Address_sortred_file.csv')
    for file in input_files:
        split_files(file)
        os.remove(file)
        
    lbl_input = glob.glob('BHANDARA_BATCH_ID_*.txt') 
    for f in lbl_input:
        total_qty_box = len(lbl_input)
        i+=1
        label(f,total_qty_box,i)
        os.remove(f)
    
    previous_bin_num = ''
    with open('xxBHANDARA_SCANNING_FILE.csv','r') as infile:
        contents = infile.readlines()
        srno = 0
        size = len(contents)
        for line in contents:
            bin_no = line.split('|')[-1].strip()
            srno += 1
            if bin_no != previous_bin_num:
                srno = 1
            with open(f'xBHANDARA_SCANNING_FILE_{bin_no}_{date_1}.txt','a') as lbl_out:
                lbl_out.write(str(srno).zfill(5)+'|'+line[:-8]+'\n')
                previous_bin_num = bin_no




    for f in glob.glob('xBHANDARA_SCANNING_FILE_*.txt'):
        with open(f ,'r') as infile:
            contents = infile.readlines()
            size = len(contents)
            srno = 0
            with open(f[1:-4]+'_'+str(size).zfill(4)+'.csv','a') as outfile:
                outfile.write('Sr. No.|Sequence Number|CARD NO.|Courier Name|Invoice_Number|BOX_NO|Branch_code\n')
                outfile.writelines(contents)
            
    with open('xxBHANDARA_LABEL_FILE.csv','r') as infile:
        contents = infile.readlines()
        sr_no = 0
        with open(f'BHANDARA_LABEL_FILE_{date_1}_{str(total_qty_lot).zfill(4)}.csv','a') as lbl_out:
            lbl_out.write('S No|Branch Name|Branch ID|First Card Id Number|Last Card Id Number|Branch/Box Qty.|Box Qty.|File Name|File Qty.|Data Qty.|Name|Add1|Add2|City|State|Pin code|Phone no|AWB no|Routing code\n')
            for line in contents:
                sr_no += 1  
                lbl_out.write(str(sr_no).zfill(5)+'|'+line)



    with open('xxBHANDARA_LABEL_SCANNING_FILE.csv','r') as infile:
        contents = infile.readlines()
        srno = 0
        with open(f'BHANDARA_LABEL_KIT_FILE_{date_1}_{str(total_qty_lot).zfill(4)}.csv','a') as lbl_out:
            
            lbl_out.write('Sr. No.|Sequence Number|CARD NO.|Courier Name|Invoice_Number|BOX_NO|Branch_code\n')
            for line in contents:
                srno += 1  
                lbl_out.write(str(srno).zfill(5)+'|'+line)
    
    with open('xxBHANDARA_PACKAGING_FILE.csv','r') as infile:
        contents = infile.readlines()
        sr_no = 0
        with open(f'BHANDARA_PACKAGING_FILE_{date_1}_{str(total_qty_lot).zfill(4)}.csv','a') as mis_out:
            mis_out.write('Sr No.|Card Id.|Cust Card No.|Branch ID|Account No.|Branch Name|Box No.\n')
            for line in contents:
                sr_no += 1  
                mis_out.write(str(sr_no).zfill(5)+'|'+line)
    
    excel_inputs = glob.glob('BHANDARA_PACKAGING_FILE_*.csv') 
    for xl in excel_inputs:
        Excel_converter(xl)
        os.remove(xl)
        
    input_files = glob.glob('BHANDARA_PACKAGING_FILE_*.xlsx')
    
    output_file = os.getcwd()+"/BHANDARA_PACKAGING_FILE_.pdf"  # Update with desired PDF output path
    for file in input_files:
        if os.path.exists(file):
            generate_pdf_from_excel(file, output_file)
        else:
            print(f"Input file not found: {file}")    
        
    
    with open('BHANDARA_BANK_File_count.txt', 'w') as f:
        print(occurrences, file=f) 
    # Read in the file
    with open('BHANDARA_BANK_File_count.txt', 'r') as file :
      filedata = file.read()
    # Replace the target string
    filedata = filedata.replace(',', '\n')
    filedata = filedata.replace('Counter({', '')
    filedata = filedata.replace('})', '')
    
    
    with open('C:/AUTO-PROCESS-CONFIG/batch_series', 'r') as fin:
        batch_series_data = fin.read().splitlines(True)
        batch_series = batch_series_data[0:1]
    fin.close()
    with open('C:/AUTO-PROCESS-CONFIG/batch_series', 'w') as fout:
        fout.writelines(batch_series_data[1:])
    fout.close()
    listToStr = ' '.join([str(elem) for elem in batch_series])
    batch_number = listToStr[0:5]
    
    file_count = '.'
    os.chdir(file_count)
    names={}
    total_qty = []
    
    
    
    
    
    files = glob.glob('*.txtt')
    for f in files:
        with open(f,'r') as emb_out, open('xBHANDARA_FPA_SCANNING_FILE.csv','a') as fpa_out:
            contents =  emb_out.readlines()
            size = len(contents)
            
            for line in contents:
                card_number = line.split(',')[1].strip()
                bin_num = line[7:14].replace(' ', '')
                # emb_out.write(line)
                fpa_out.write(card_number[-7:].replace(' ','')+'|'+f+'|'+date_1+'\n')
       
    with open('xBHANDARA_FPA_SCANNING_FILE.csv','r') as infile:
        contents = infile.readlines()
        srno = 0
        size = len(contents)
        with open(f'BHANDARA_FPA_SCANNING_FILE_{date_1}_{str(size).zfill(4)}.csv','a') as lbl_out:
            lbl_out.write('Sr. No.|LAST SIX CARD NO.|FILENAME|Date\n')
            for line in contents:
                srno += 1  
                lbl_out.write(str(srno).zfill(5)+'|'+line)
    
    for fn in glob.glob('*.txtt*'):
        with open(fn) as f:
            names[fn]=sum(1 for file_count in f if file_count.strip() and not file_count.startswith('~'))       
    
    
    currentTime = time.strftime(".%H.%M.%S", Time)
    currenthour = time.strftime("%H", Time)
    
    
    
    ptx = PrettyTable()
    ptx.field_names = ["Bin", "Artwork No.","JB No.", "Emboss Filename","Quantity", "Job setup", "Printing method - Front/Back"]
    ptx1 = PrettyTable()
    ptx1.field_names = ["Emboss Filename","Supervisor Name","Supervisor Signature", "  Date  ","  Time  ","   Remark(if any)   "]
    
    for file in glob.glob("*.txtt"):
        bin_no=file.split('_')[3].rstrip()
        pro_name = '_'.join(file.split('_')[4:9]).rstrip()
        with open(file) as f:
            qty2=len(f.readlines())
        with open('CONFIG/job_setup.csv','r') as job,open('CONFIG/batchcard_file.csv','r') as batch_job:
            art = ''   
            jset= ''
            m_r = ''
            datas = batch_job.readlines()
            for j in job.readlines()[1:]:
                bin_no_1 = j.split(',')[0].rstrip()
                
                if bin_no_1 == bin_no :
                    art=j.split(',')[1].rstrip()   
                    jset=j.split(',')[2].rstrip()
                    m_r=j.split(',')[3].rstrip()
                    break
            
            for data in datas[1:]:
                value = data.split(',')
                batch_bin_num = value[0].rstrip()
                
                if batch_bin_num == bin_no :
                    printing_front = value[4].rstrip()
                    job_setup_batch = value[1].rstrip()
                    ribbon_front = value[5].rstrip()
                    printing_back = value[6].rstrip()
                    ribbon_back = value[7].rstrip()
                    batch_product = value[2].rstrip()
                    batch_art = value[3].rstrip()
                    
                    with open('BHANDARA_BANK_BATCHCARD_FILE.csv','a') as out_file:
                        out_file.write(bin_no+'|'+job_setup_batch+'|'+file+'|'+str(qty2).rjust(4,'0')+'|'+batch_product+'|'+batch_art+'|'+printing_front+'|'+ribbon_front+'|'+printing_back+'|'+ribbon_back+'|'+'NEW ISSUE'+'\n')
                    break
                
                
            
            
        ptx.add_row([bin_no, art ,"", file,str(qty2).rjust(4,'0'), jset, m_r])
        ptx.align = "c"
        ptd=ptx.get_string()
        
        ptx1.add_row([file,"","","","",""])
        ptx1.add_row(["","","","","",""])
        ptx1.align = "c"
        ptd1=ptx1.get_string()
        total_qty.append(qty2)
    datedmy1 = str(x)[8:10]+str(x)[5:7]+str(x)[0:4]
    datedmy2 = str(x)[8:10]+'-'+str(x)[5:7]+'-'+str(x)[0:4]
    datedmy = str(x)[8:10]+'-'+str(x)[5:7]+'-'+str(x)[0:4]
    
    with open('BHANDARA_BANK_BATCHCARD_'+str(x)+'_'+str(total_qty_lot).zfill(4)+'.dat', 'w') as file:
      file.write('                                             '+'Banking Data Preparation Batch Card'+'                       '+'Date: '+datedmy+currentTime+'\n')
      file.write('                                                       BHANDARA_BANK-'+str(x)[8:10]+str(x)[5:7]+str(x)[0:4]+'-'+batch_number+'                         '+'BHANDARA_BANK_PROJECT(DI)\n\n')
      #file.write('EMBOSS FILE NAME: BHANDARA_EMBOSS_'+str(x)+'_'+str(qty)+'.txtt\n')
      file.write(str(ptd))
      file.write('\n')
      file.write('TOTAL BATCH QUANTITY:'+str(total_qty_lot).zfill(4)+'\n\n')
      file.write('Data upload on Machine\n')
      file.write(str(ptd1))
      file.write('\n\n\n\n\n\n\n\n\n\n\n\n\n')
      file.write('PRP: 20.1                                        Rev No: 3.3                                       Date: 01-Feb-25\n')
      file.write('SEC-3: INTERNAL                                  Owner: Quality Control                            Status: Issued\n')
      file.write('                                                  Page: 1 of 1')
    
    with open('BHANDARA_BANK_File_count_'+str(x)+'.txt', 'w') as file:
      file.write('Total Quantity:'+str(total_qty_lot).zfill(4)+'\n\n')
      file.write("Files Processed\n")
      file.write(filedata)
    
    with open('C:/AUTO-PROCESS-CONFIG/BHANDARA_BANK_PROCESSING_LOG.log', 'a') as file:
    
      file.write("Files Processed\n") 
      date_and_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
      file.write(date_and_time+'\n'+filedata)
      file.write("Files Deleted\n")
      file.write(date_and_time+'\n'+filedata+'\n\n')
    
    with open('BHANDARA_BANK_BATCHCARD_FILE.csv','r') as in_file, open('BHANDARA_BANK_BATCHCARD_'+str(x)[8:10]+str(x)[5:7]+str(x)[0:4]+'_'+str(total_qty_lot).zfill(4)+'.TXT','a') as output_file:
        contents = in_file.readlines()
        output_file.write('Sr no.|BIN|JOB SETUP|FILE NAME|Quantity|LOGO|ARTWORK NO.|Printing Front|Ribbon Front|Printing Back|Ribbon Back|Card Action\n')
        sn = 0
        for line in contents:
            sn+=1
            output_file.write(str(sn).zfill(5)+'|'+line)
            
    df = pd.read_csv('BHANDARA_BANK_BATCHCARD_'+str(x)[8:10]+str(x)[5:7]+str(x)[0:4]+'_'+str(total_qty_lot).zfill(4)+'.TXT',encoding= 'unicode_escape',sep='|', dtype=object)
    df.to_excel('BHANDARA_BANK_BATCHCARD_'+str(date_1)+'_'+str(total_qty_lot).zfill(4)+'.xlsx', 'Sheet1', index=False)
    os.remove('BHANDARA_BANK_BATCHCARD_FILE.csv')
    
    path = ('.')
    ext = "txtt"
    
    for f in os.listdir(path):
        fpath = os.path.join(path, f)
    
        if os.path.isfile(fpath) and fpath.endswith(ext):
            time = datetime.fromtimestamp(os.path.getctime(fpath)).strftime("%d-%m-%Y-%H%M%S---"+str(total_qty_lot).zfill(4))
            name=f'BHANDARA_BANK_EMBO_{batch_number}_'+time
            os.makedirs(os.path.join(path, name), exist_ok=True)
            os.replace(fpath, os.path.join(path, name, f))
            
       
    import shutil       
    
    fromDirectory = os.getcwd()+'/'
    aaa = fromDirectory+name
    toDirectory = "A:/Sdrive/LIVE/BHANDARA_BANK/"
    bbb = toDirectory+'/'+name
    try:
        shutil.copytree(aaa,bbb)
    except Exception as e:
        print('Error while copying embo file to A:/Sdrive/LIVE/BHANDARA_BANK/ drive error msf : '+str(e))
    
    
    
    
    
    
    
    from fpdf import FPDF
    pdf = FPDF()
    pdf.add_page('L')
    pdf.set_font("Courier", size=6.3)
    f = open('BHANDARA_BANK_BATCHCARD_'+str(x)+'_'+str(total_qty_lot).zfill(4)+'.dat', "r")
    #f
    for xpdf in f:
        pdf.cell(50, 5, txt = xpdf, ln = 1, align = 'L')
        pdf.image('C:/AUTO-PROCESS-CONFIG/colorplast_logo.png', x = 10, y = 5, w = 50, h = 10, type = '', link = '')  
    pdf.output('BHANDARA_BANK_BATCHCARD_'+datedmy+'_'+batch_number+'_'+str(total_qty_lot).zfill(4)+'.pdf')
    f.close()
    
    
    pdfl = FPDF()
    pdfl.add_page('L')
    pdfl.set_font("Courier", size=10)
    flog = open('C:/AUTO-PROCESS-CONFIG/BHANDARA_BANK_PROCESSING_LOG.log', "r")
    #flog
    for xlog in flog:
        pdfl.cell(50, 5, txt = xlog, ln = 1, align = 'L')
         
    pdfl.output('BHANDARA_BANK_PROCESSING_LOG'+'.pdf')
    pdfl.close()
    
    from prettytable import PrettyTable
    
    ptx = PrettyTable()
    ptx.field_names = (["Client Name", "File Name","File Deletion Date", "Dispatch Date","Data Admin. Name", "Data Admin. Sign.","IT Person Name", "IT Person Sign."])
    
    
    
    for key,value in names.items():
        
        ptx.add_row(["BHANDARA Bank", key, "", "", "Data Team", "", "IT Team", ""'\n'])
    ptx.align = "c"
    ptd=ptx.get_string()
    
    with open('BHANDARA_BANK_DELETION_LOG_'+datedmy+'.dat', 'w') as file:
      file.write('\n\n                                                                        SEC-IS   |   02.02\n')
      file.write('                                                                        Rev.No.  |   3.0                                                   '+'BHANDARA PROJECT\n')
      file.write('                                                                        Date :   |   15-Jul-18\n')
      file.write('                                                                '+'( ***** DATA DELETION LOG ***** )''\n\n\n\n') 
      file.write('|  Data Receiving Date - '+datedmy+'  |  '+'  Data Servier - DPP Server/MX Machine  |   '+'  Batch Qty. - '+str(total_qty_lot).zfill(4)+'   |  '+'  Batch No. - SBM-'+str(x)[8:10]+'-'+str(x)[5:7]+'-'+str(x)[0:4]+'--'+batch_number+'  |  '+'Status -             '+'|''\n\n\n')
      file.write(str(ptd))
      file.write('\n\n\n\n')
      #file.write('BATCH NO. OF PM TOOL FOR THIS BATCHCARD :_______________\n')
    from fpdf import FPDF
    
    pdf = FPDF()
    pdf.add_page('L')
    pdf.set_font("Courier", size=6.4)
    f = open('BHANDARA_BANK_DELETION_LOG_'+datedmy+'.dat', "r")
    #f
    for xpdf in f:
        pdf.cell(50, 5, txt = xpdf, ln = 1, align = 'L')
        pdf.image('C:/AUTO-PROCESS-CONFIG/colorplast_logo.png', x = 10, y = 20, w = 50, h = 10, type = '', link = '')  
    pdf.output('BHANDARA_BANK_DELETION_LOG_'+datedmy+'.pdf')
    f.close()
    
    # ---------------------------------DELETION LOG ENDS ---------------------------
    
    del_file1 = glob.glob(f'*{filename}*')
    for de in del_file1:   
        os.remove(de)
    
    del_filetxt = glob.glob("*.txt")
    for de in del_filetxt:   
        os.remove(de)
    
    
    
    destination_folder = "C:/Config/Batchcard"
    os.makedirs(destination_folder, exist_ok=True)
    files = glob.glob('*BATCHCARD_*.dat') + glob.glob('*DELETION_LOG_*.dat')
    for file in files:
        destination_path = os.path.join(destination_folder, os.path.basename(file))
        shutil.move(file, destination_path)    
    
    
    
    del_inputs = glob.glob('*xB*')  + glob.glob('*.txt') + glob.glob('*TBDCC*') #+ glob.glob('*BDCCB_*')
    for dl_file in del_inputs:
        os.remove(dl_file)
        
        
    excel_inputs = glob.glob('BHANDARA*.csv') 
    for xl in excel_inputs:
        Excel_converter(xl)
        os.remove(xl)
        
    # inp_file = glob.glob('*Packaging_*.xlsx')
    # for file in inp_file: 
    #     # Read Excel file
    #     df = pd.read_excel(file, engine='openpyxl', dtype=object)
        
    #     # Initialize PDF canvas
    #     c = canvas.Canvas("BHANDARA_PackingList.pdf", pagesize=letter)
    #     width, height = letter
        
    #     # Set column number (1-indexed)
    #     column_number = 8  # Assuming you want to use column 4 (1-indexed)
        
    #     # Header text (assuming headers are known)
    #     headers = ["Sr No.","Name", "Cust Card No.", "Branch Name", "Account No.",  "Branch ID","Card Id.","Box No."]
        
    # draw_header()
    
    # draw_content()
    
    # c.save()
    
    # input_files= glob.glob('*.xlsx')
    # for f in input_files:
    #     # f.replace(f'{date_1}',f'{date_1}_{batch_number}')
    #     new_name = f.replace(f'{date_1}', f'{date_1}_{batch_number}')
        
        
    #     os.rename(f, new_name)
    
    
        
    current_directory = os.getcwd()
    excel_files = glob.glob(os.path.join(current_directory, '*FF*.xlsx')) + glob.glob(os.path.join(current_directory, '*Scanning*.xlsx'))
    
    if excel_files:
        for excel_file in excel_files:
            convert_to_read_only(excel_file)
    
    
    destination_folder = "C:/Config/Batchcard"
    os.makedirs(destination_folder, exist_ok=True)
    files = glob.glob('*BATCHCARD_*.dat') + glob.glob('*DELETION_LOG_*.dat')
    for file in files:
        destination_path = os.path.join(destination_folder, os.path.basename(file))
        shutil.move(file, destination_path)    
    
    from distutils.dir_util import copy_tree
    path = ('.')
    dt1=datetime.now().strftime("%Y-%m-%d-%H%M%S---")
    
    
    for f in os.listdir(path):
        fpath = os.path.join(path, f)
        if os.path.isfile(fpath) and (fpath.endswith('xlsx') or fpath.endswith('pdf') or fpath.endswith('csv') or fpath.endswith('txt')):
            # time = datetime.fromtimestamp(os.path.getctime(fpath)).strftime("BOM_FIS-"+"%d-%m-%Y---"+str(qty).rjust(4,'0'))dat
            time = datetime.fromtimestamp(os.path.getctime(fpath)).strftime(f"BHANDARA_BANK_FF_MIS_{batch_number}_"+dt1+str(total_qty_lot).zfill(4))
            os.makedirs(os.path.join(path, time), exist_ok=True)
            os.replace(fpath, os.path.join(path, time, f))
    
    with open('CONFIG/Output_file_location.csv','r') as file:
        content = file.readlines()
        folder_location = content[0].split(',')[1].strip()
        
    os.chdir(path) 
    fromDirectory = os.getcwd()
    aaa = fromDirectory+'/'+time
    toDirectory = folder_location
    bbb = toDirectory+time
    copy_tree(aaa, bbb)
    closeInput = input("Press ENTER to exit")
    print ("Closing...")










